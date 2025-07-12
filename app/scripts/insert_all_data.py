import psycopg2
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash
from datetime import datetime
from decimal import Decimal

def safe_float(val):
    try:
        if isinstance(val, str):
            val = val.replace("RM", "").replace(",", "").replace("%", "").strip()
        return float(val)
    except:
        return 0.0
    
def safe_cell(row, idx):
    try:
        return str(row[idx]).strip() if idx < len(row) and row[idx] else ""
    except:
        return ""

def safe_float_cell(row, idx):
    try:
        val = row[idx] if idx < len(row) else None
        if isinstance(val, str):
            val = val.replace("RM", "").replace(",", "").replace("%", "").strip()
        return float(val or 0)
    except:
        return 0.0

# Connect to DB
conn = psycopg2.connect(
    dbname="propdb",
    user="propuser",
    password="proppass123",
    host="localhost",
    port="5432"
)
cur = conn.cursor()

# ------------------------------------
# INSERT DEFAULT USERS
# ------------------------------------
users = [
    {"name": "Mohd Anuar", "email": "mohdanuar.admkedah15@gmail.com", "password": "anuar123", "role": "Admin", "team": "HQ"},
    {"name": "Shaza Admin", "email": "shaza.proper@gmail.com", "password": "shaza123", "role": "Admin", "team": "HQ"},
    {"name": "Eusoff Admin", "email": "eusoff@proper.com", "password": "eusoff123", "role": "Admin", "team": "HQ"},
    {"name": "Azimi Leader", "email": "azimi@proper.com", "password": "azimi123", "role": "Leader", "team": "KEDAH"},
    {"name": "Mizan Agent", "email": "mizan@proper.com", "password": "mizan123", "role": "Agent", "team": "KEDAH"},
    {"name": "Azmira Admin", "email": "azmira@proper.com", "password": "azmira123", "role": "Admin", "team": "HQ"}
]

user_count = 0
for user in users:
    cur.execute("SELECT * FROM users WHERE email = %s", (user["email"],))
    if not cur.fetchone():
        hashed_pw = generate_password_hash(user["password"])
        cur.execute("""
            INSERT INTO users (name, email, password_hash, role, team)
            VALUES (%s, %s, %s, %s, %s)
        """, (user["name"], user["email"], hashed_pw, user["role"], user["team"]))
        print(f"✅ Inserted user: {user['email']}")
        user_count += 1
    else:
        print(f"⚠️ Already exists: {user['email']}")

# ------------------------------------
# IMPORT AGENTS FROM EXCEL
# ------------------------------------
cur.execute("DELETE FROM users WHERE role = 'Agent'")
print("🧹 Cleared existing Agents.")

wb = load_workbook("data/AGENT 2025.xlsx", data_only=True)
ws = wb.active
agent_count = 0

for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[1] or row[1] == "Nama":
        continue

    name = row[1]
    ren_no = row[2] or ""
    phone = row[3] or ""
    nric = row[4] or ""
    email = row[5]
    tiering_pct = safe_float(row[6])
    position = row[7] or ""
    notes = row[8] or ""
    password_hash = generate_password_hash("agent123")

    if not email:
        with open("skipped_agents.txt", "a") as f:
            f.write(f"{name}, {phone}, {ren_no}\n")
        continue

    try:
        cur.execute("""
            INSERT INTO users (name, email, password_hash, role, team, phone, ren_no, nric,
                               tiering_pct, position, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (name, email, password_hash, "Agent", position, phone, ren_no, nric, tiering_pct, position, notes))
        print(f"✅ Inserted agent: {name}")
        agent_count += 1
    except Exception as e:
        print(f"❌ Error inserting {name}: {e}")
        conn.rollback()

# ------------------------------------
# IMPORT CASES FROM EXCEL
# ------------------------------------

from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook("data/LIST CASE APRIL.xlsx", data_only=True)
ws = wb.active
cases_count = 0

for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
    if not row or not row[1] or "NO CASE" in str(row[1]):
        continue

    row = list(row) + [""] * 20  

    try:
        number            = int(float(row[0])) if row[0] else None
        case_no           = safe_cell(row, 1)
        invoice_no        = safe_cell(row, 2)
        eform_id          = safe_cell(row, 3)
        agent_name        = safe_cell(row, 4).replace("PA :", "").replace("VA :", "").strip().lower()
        client_name       = safe_cell(row, 5)
        description       = safe_cell(row, 6)
        purchase_price    = safe_float_cell(row, 7)
        amount            = safe_float_cell(row, 8)
        ed_paid           = safe_float_cell(row, 9)
        ed_pending        = safe_cell(row, 10)
        tax               = safe_float_cell(row, 12)
        total_amount      = safe_float_cell(row, 13)
        commission_pct    = safe_float_cell(row, 14)
        commission_total  = safe_float_cell(row, 15)
        override_leader   = safe_float_cell(row, 16)
        override_hoa      = safe_float_cell(row, 17)
        profit_proper     = safe_float_cell(row, 18)
        case_status       = safe_cell(row, 19) or ("SETTLE" if commission_total > 0 and profit_proper > 0 else "IN PROGRESS")
        date_created      = "2025-04-15"

        # Get agent_id
        cur.execute("SELECT user_id FROM users WHERE LOWER(name) LIKE %s", (f"%{agent_name}%",))
        agent_result = cur.fetchone()
        agent_id = agent_result[0] if agent_result else None

        if not agent_id:
            print(f"⚠️ Agent '{agent_name}' not found – inserting case with NULL agent_id.")
            cur.execute("SELECT name FROM users WHERE LOWER(name) LIKE %s", (f"%{agent_name.split()[0]}%",))
            suggestions = [r[0] for r in cur.fetchall()]
            if suggestions:
                print(f"🔍 Suggestions: {suggestions}")

        # INSERT CASE INTO DB
        cur.execute("""
            INSERT INTO cases (
                number, case_no, invoice_no, eform_id, agent_id, leader_id,
                agent_name, leader_name, client_name, property_address,
                description, case_details, purchase_price, amount, fee_pct, commission_pct,
                commission_total, override_leader_amt, override_hoa_amt, profit_proper,
                ed_paid, ed_pending, tax, total_amount,
                reference_no, registration_no, mode_of_payment, in_part_payment_of,
                date_created, payment_status, case_status
            ) VALUES (%s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s)
        """, (
            number, case_no, invoice_no, eform_id, agent_id, None,
            agent_name, "", client_name, "",
            description, "", purchase_price, amount, 0.0, commission_pct,
            commission_total, override_leader, override_hoa, profit_proper, ed_paid, ed_pending, tax, total_amount,
            "", "", "", "",
            date_created, '', case_status
        ))

        print(f"✅ Inserted case row {i}: {case_no}")
        cases_count += 1

    except Exception as e:
        print(f"❌ Error inserting row {i}: {e}")
        conn.rollback()

# ------------------------------------
# GENERATE INVOICES
# ------------------------------------
cur.execute("""
    SELECT case_id, agent_id, leader_id, total_amount, commission_total, case_no
    FROM cases
    WHERE agent_id IS NOT NULL
""")
cases = cur.fetchall()

invoice_count = 0
for case in cases:
    case_id, agent_id, leader_id, total_amount, commission_total, case_no = case

    agent_comm = round(Decimal(commission_total) * Decimal("0.70"), 2)
    leader_comm = round(Decimal(commission_total) * Decimal("0.15"), 2) if leader_id else Decimal("0.00")
    admin_comm = round(Decimal(commission_total) - (agent_comm + leader_comm), 2)

    invoice_no = f"INV-{str(case_id).zfill(4)}"
    date_issued = datetime.today().strftime("%Y-%m-%d")

    try:
        cur.execute("""
            INSERT INTO invoices (
                invoice_no, case_id, agent_id, leader_id,
                agent_commission, leader_commission, admin_commission,
                commission_pct, commission_total, tax, total_amount,
                date_issued, remarks
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            invoice_no, case_id, agent_id, leader_id,
            agent_comm, leader_comm, admin_comm,
            None, commission_total, None, total_amount,
            date_issued, f"Auto-generated for {case_no}"
        ))
        invoice_count += 1
    except Exception as e:
        print(f"❌ Invoice error for {case_no}: {e}")
        conn.rollback()

# ------------------------------------
# DUMMY LEDGER ENTRY
# ------------------------------------
cur.execute("SELECT invoice_id FROM invoices LIMIT 1")
first_invoice = cur.fetchone()
if first_invoice:
    cur.execute("""
        INSERT INTO ledger (invoice_id, amount_paid, date_paid, payment_method, payment_note)
        VALUES (%s, %s, %s, %s, %s)
    """, (first_invoice[0], 500.00, "2025-06-06", "Online Transfer", "Initial deposit"))

# ------------------------------------
# TEST REGISTRATION REQUEST
# ------------------------------------
cur.execute("""
    INSERT INTO registration_requests (
        name, email, phone, team, role, password_hash, nric, dob
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
""", (
    "New User", "newuser@proper.com", "0123456789", "HQ", "Agent",
    generate_password_hash("test123"), "900101-14-5678", "1990-01-01"
))

# Finalize
conn.commit()
cur.close()
conn.close()
print(f"🎯 Done: Inserted {user_count} users, {agent_count} agents, {cases_count} cases & {invoice_count} invoices.")